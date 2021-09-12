#required openpyxl installed
from openpyxl import load_workbook,Workbook
class Sheet:
	def __init__(self,sheet_name,data_matrix):
		self.sheet_name=sheet_name
		self.data_matrix=data_matrix
    
def read_excel(file_name):
	#return a list of Sheet
	workbook=load_workbook(filename=file_name)
	ans=[]
	for sheet in workbook:
		data=[]
		for row in sheet.rows:
			data=data+[list(filter(lambda x:x!=None,[cell.value for cell in row]))]
		data=list(filter(lambda x:x!=[],data))
		ans=ans+[Sheet(sheet.title,data)]
	return ans
  
def write_excel(file_name,sheets):
	#Each element in sheets should be Sheet
	workbook=Workbook()
	del workbook[workbook.active.title]
	for sheet in sheets:
		worksheet=workbook.create_sheet(sheet.sheet_name)
		data=sheet.data_matrix
		for r in range(0,len(data)):
			for c in range(0,len(data[r])):
				worksheet.cell(column=c+1,row=r+1,value=data[r][c])
	workbook.save(filename=file_name)
